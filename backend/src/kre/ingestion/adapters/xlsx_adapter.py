from pathlib import Path

from openpyxl import load_workbook

from kre.models import Chunk


def parse(path: Path, document_id: str) -> list[Chunk]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    chunks: list[Chunk] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                if not text:
                    continue
                chunks.append(Chunk(
                    id=f"{document_id}:xlsx:{sheet.title}:{cell.coordinate}",
                    document_id=document_id, source_format="xlsx", text=text,
                    element_type="cell", section_path=(sheet.title,),
                    location_reference=f"Sheet: {sheet.title}, Cell: {cell.coordinate}",
                    metadata={"sheet": sheet.title, "coordinate": cell.coordinate},
                ))
    return chunks
